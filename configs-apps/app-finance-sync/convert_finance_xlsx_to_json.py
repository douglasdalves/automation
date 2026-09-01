#!/usr/bin/env python3
"""Convert a finance workbook into the JSON expected by dashboard.html.

Expected Excel layout:
- One sheet with columns: mes, receita, contas_mensais, extras, despesas, saldo, reserva, investimentos
- Optional second sheet with columns: item, total (for categorias)
- Optional third sheet with columns: item, valor (for investimentos_recentes.itens)
- Optional cell/row with the month label for investimentos_recentes.mes

If the file is not shaped like that, the script will print a helpful error and exit.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import tempfile
import unicodedata
from pathlib import Path
from typing import Any, Iterable
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Dependency missing: install with `pip install -r requirements.txt`") from exc


def normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def parse_decimal(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace("R$", "").replace(" ", "")
    if text in ("", "-", "—"):
        return 0.0

    if "." in text and "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text and "." not in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_key(value: Any) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def header_aliases() -> dict[str, set[str]]:
    return {
        "mes": {"mes", "meses", "mês"},
        "receita": {"receita", "receitas"},
        "contas_mensais": {"contas_mensais", "contas_fixas", "contasfixas", "contas_fixas", "fixas"},
        "extras": {"extras", "extra"},
        "despesas": {"despesas", "despesa", "gastos"},
        "saldo": {"saldo", "saldo_geral", "saldogerald"},
        "reserva": {"reserva", "reservas"},
        "investimentos": {"investimentos", "investimento"},
        "item": {"item", "categoria", "classe", "nome"},
        "total": {"total", "valor_total", "valor", "valor_total_mensal"},
        "valor": {"valor", "total", "total_mensal", "valor_total"},
    }


def header_index(headers: Iterable[Any], desired: Iterable[str]) -> dict[str, int]:
    normalized = [normalize_key(h) for h in headers]
    aliases = header_aliases()
    result: dict[str, int] = {}
    for target in desired:
        accepted = {normalize_key(v) for v in aliases.get(target, {target})} | {target}
        for idx, header in enumerate(normalized):
            if header in accepted:
                result[target] = idx
                break
    return result


def first_sheet_with_header(workbook, candidates: list[str]):
    for ws in workbook.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [normalize_key(v) for v in rows[0]]
        normalized_candidates = {normalize_key(c) for c in candidates}
        if any(c in headers for c in normalized_candidates):
            return ws
    return None


def normalize_month_label(sheet_name: str, year_hint: int | None = None) -> str:
    text = str(sheet_name or "").strip()
    if not text:
        return text

    normalized = normalize_key(text)
    month_names = {
        "janeiro": 1, "jan": 1, "january": 1,
        "fevereiro": 2, "fev": 2, "feb": 2, "february": 2,
        "marco": 3, "mar": 3, "march": 3,
        "abril": 4, "apr": 4, "april": 4,
        "maio": 5, "may": 5,
        "junho": 6, "jun": 6, "june": 6,
        "julho": 7, "jul": 7, "july": 7,
        "agosto": 8, "ago": 8, "aug": 8, "august": 8,
        "setembro": 9, "set": 9, "sep": 9, "september": 9,
        "outubro": 10, "out": 10, "oct": 10, "october": 10,
        "novembro": 11, "nov": 11, "november": 11,
        "dezembro": 12, "dez": 12, "dec": 12, "december": 12,
    }

    month_number = None
    for key, value in month_names.items():
        if key in normalized:
            month_number = value
            break

    if month_number is None:
        return text

    match = re.search(r"(19|20)\d{2}|\b(\d{2})\b", text)
    if match:
        year_raw = match.group(1) or match.group(2)
        year = int(f"20{year_raw}") if len(year_raw) == 2 else int(year_raw)
    elif year_hint is not None:
        year = year_hint if month_number != 12 else year_hint - 1
    else:
        year = 2025

    month_abbr = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"][month_number - 1]
    return f"{month_abbr}/{str(year)[-2:]}"


def _infer_year_hint(workbook) -> int | None:
    years = []
    for ws in workbook.worksheets:
        match = re.search(r"(19|20)\d{2}|\b(\d{2})\b", ws.title)
        if not match:
            continue
        year_raw = match.group(1) or match.group(2)
        if len(year_raw) == 2:
            years.append(2000 + int(year_raw))
        else:
            years.append(int(year_raw))
    return max(years) if years else None


def _sum_numeric_values_after_label(rows, labels: set[str], target_index: int | None = None):
    total = 0.0
    for idx, row in enumerate(rows[:-1]):
        row_labels = {normalize_key(v) for v in row}
        if not row_labels.intersection(labels):
            continue
        next_row = rows[idx + 1]
        if target_index is not None:
            if target_index < len(next_row):
                total += parse_decimal(next_row[target_index])
        else:
            for value in next_row:
                if isinstance(value, (int, float)):
                    total += parse_decimal(value)
    return total


def load_month_rows(workbook):
    year_hint = _infer_year_hint(workbook)
    months = []

    for ws in workbook.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        first_text = " ".join(str(v or "") for v in rows[0][:10]).lower()
        if "saldos" not in first_text and "entrada mês" not in first_text and "entrada mes" not in first_text:
            continue

        revenue = 0.0
        for idx, row in enumerate(rows[:-1]):
            row_labels = {normalize_key(v) for v in row}
            if not row_labels.intersection({"recebido", "adicional", "descontos"}):
                continue
            next_row = rows[idx + 1]
            if len(next_row) > 1 and isinstance(next_row[1], (int, float)):
                revenue += parse_decimal(next_row[1])

        contas_mensais = _sum_numeric_values_after_label(rows, {"mensais"}, 3)
        extras = _sum_numeric_values_after_label(rows, {"extras", "gastos"}, 3)
        saldo = _sum_numeric_values_after_label(rows, {"saldo_geral", "saldo", "resultados"}, 4)
        reserva = _sum_numeric_values_after_label(rows, {"reserva"}, 4)
        investimentos = 0.0
        for idx, row in enumerate(rows[:-2]):
            row_labels = {normalize_key(v) for v in row}
            if "investimentos" not in row_labels:
                continue
            future_row = rows[idx + 2]
            for value in future_row:
                if isinstance(value, (int, float)):
                    investimentos = parse_decimal(value)
                    break

        if revenue == 0 and contas_mensais == 0 and saldo == 0:
            continue

        month_data = {
            "label": normalize_month_label(ws.title, year_hint),
            "receita": revenue,
            "contas_mensais": contas_mensais,
            "extras": extras,
            "despesas": contas_mensais + extras,
            "saldo": saldo,
            "reserva": reserva,
            "investimentos": investimentos,
        }
        months.append(month_data)

    if not months:
        raise ValueError("Nenhuma aba de meses foi identificada. Verifique o nome da aba ou a estrutura da planilha.")

    labels = [m["label"] for m in months]
    return labels, months


def load_categories(workbook):
    totals: dict[str, float] = {}
    for ws in workbook.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        for row in rows:
            if len(row) < 9:
                continue
            item = row[7]
            value = row[8] if len(row) > 8 else None
            text = str(item or "").strip()
            if not text or text.lower() in {"item", "total", "status", "valor"}:
                continue
            if isinstance(value, (int, float)):
                totals[text] = totals.get(text, 0.0) + parse_decimal(value)

    result = [{"item": item, "total": total} for item, total in sorted(totals.items(), key=lambda kv: kv[1], reverse=True)]
    return result


def load_investment_recent(workbook):
    recent_label = ""
    recent_items: dict[str, float] = {}

    for ws in workbook.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        label = normalize_month_label(ws.title, _infer_year_hint(workbook))
        items: dict[str, float] = {}
        collecting = False
        for row in rows:
            row_text = " ".join(str(v or "") for v in row).lower()
            if "investimentos" in row_text:
                collecting = True
                continue
            if not collecting:
                continue
            if "contas ap" in row_text or "contas toronto" in row_text:
                break

            item = None
            item_index = None
            for idx, value in enumerate(row):
                text = str(value or "").strip()
                if not text:
                    continue
                if text.lower() in {"item", "valor", "status", "total"}:
                    continue
                if isinstance(value, str):
                    item = text
                    item_index = idx
                    break

            if item is None:
                continue

            value = None
            for candidate in row[item_index + 1:]:
                if isinstance(candidate, (int, float)):
                    value = candidate
                    break
            if value is not None:
                items[item] = parse_decimal(value)

        if items:
            recent_label = label
            recent_items = items

    if not recent_items:
        return {"mes": "", "itens": []}

    clean_items = [{"item": item, "valor": value} for item, value in recent_items.items()]
    return {"mes": recent_label, "itens": clean_items}


def build_payload(workbook):
    labels, months = load_month_rows(workbook)
    categories = load_categories(workbook)
    recent = load_investment_recent(workbook)

    if not recent.get("itens"):
        recent = {
            "mes": labels[-1] if labels else "",
            "itens": [
                {"item": "cripto", "valor": 0},
                {"item": "Invest ações", "valor": 0},
                {"item": "Dollar", "valor": 0},
                {"item": "tesouro", "valor": 0},
                {"item": "Reservas", "valor": 0},
            ],
        }

    return {
        "labels": labels,
        "months": months,
        "categorias": categories,
        "investimentos_recentes": recent,
    }


def is_url(value: str) -> bool:
    return value.startswith(("http://", "https://", "file://"))


def append_download_param(url: str) -> str:
    if "download=1" in url:
        return url
    separator = "&" if "?" in url else "?"
    return f"{url}{separator}download=1"


def resolve_source_path(source: str) -> Path:
    if not is_url(source):
        return Path(source)

    if source.startswith("file://"):
        return Path(source[7:])

    download_url = append_download_param(source)
    request = Request(download_url, headers={"User-Agent": "Mozilla/5.0"})

    with urlopen(request, timeout=60) as response:
        content_disposition = response.headers.get("Content-Disposition", "")
        filename_match = re.search(r'filename\s*=\s*"?([^";]+)"?', content_disposition)
        suffix = ".xlsx"
        if filename_match:
            suffix = Path(filename_match.group(1)).suffix or suffix
        temp_fd, temp_path = tempfile.mkstemp(prefix="finance_", suffix=suffix)
        os.close(temp_fd)
        with open(temp_path, "wb") as file_obj:
            shutil.copyfileobj(response, file_obj)

    return Path(temp_path)


def main():
    parser = argparse.ArgumentParser(description="Converte a planilha financeira em JSON usado pelo dashboard.")
    parser.add_argument("source", help="Arquivo .xlsx local ou link compartilhado do OneDrive/Google Drive")
    parser.add_argument("--output", default="finance-data.json", help="Arquivo JSON de saída")
    args = parser.parse_args()

    source = resolve_source_path(args.source)
    if not source.exists():
        raise SystemExit(f"Arquivo não encontrado: {source}")

    wb = load_workbook(source, data_only=True)
    payload = build_payload(wb)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Arquivo gerado: {out}")


if __name__ == "__main__":
    main()
